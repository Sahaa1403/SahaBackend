from datetime import datetime, timedelta
import os
import jdatetime
import zipfile
import csv
import tempfile
from django.http import FileResponse
from django.shortcuts import get_object_or_404

from rest_framework import status
from rest_framework.response import Response
from rest_framework.views import APIView
from rest_framework.permissions import AllowAny,IsAuthenticated
from accounts.models.user import User
from search.functions.public_functions import assign_default_labels_to_kbs
from search.functions.public_functions import create_kb_process_status
from search.models import SearchData,KnowledgeBase, Label, Source, SocialMedia, KnowledgeBaseLabelUser
from search.serializers import KnowledgeBaseProcessByAiSerializer, SearchSerializer, AddKnowledgeBaseSerializer, \
    KnowledgeBaseSerializer, LabelSerializer, CreateSourceSerializer, SocialMediaProcessByAiSerializer, SourceSerializer, \
    SourceFullSerializer, SourceWithKBSerializer, EditSourceSerializer,SocialMediaSerializer,\
    KnowledgeBaseLabelUserSerializer, CreateKnowledgeBaseLabelUserSerializer, CreateSocialMedia
import requests
import logging
import time
from rest_framework.pagination import LimitOffsetPagination, PageNumberPagination
from django_filters.rest_framework import DjangoFilterBackend
from rest_framework.filters import SearchFilter, OrderingFilter
from rest_framework.generics import GenericAPIView
import json
import math
import unicodedata
import random
from django.http import HttpResponse
from io import BytesIO
import xlsxwriter
from django.db import models
from django.db import transaction
from django.db.models import Count, Q, F
import django_filters
from search.tasks import send_kb_to_ai
import json
from django.core.files.base import ContentFile
import openpyxl
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
import openpyxl
from io import BytesIO
import requests
from rest_framework.parsers import MultiPartParser
from .models import KnowledgeBaseProcessStatus, SocialMedia, Label
from django.shortcuts import get_object_or_404
from django.utils.translation import gettext as _
from uuid import uuid4

logger = logging.getLogger(__name__)



class CustomPagination(PageNumberPagination):
    page_size = 10
    page_size_query_param = 'page_size'
    max_page_size = 100

    def get_paginated_response(self, data):
        total_pages = math.ceil(self.page.paginator.count / self.get_page_size(self.request))
        return Response({
            'count': self.page.paginator.count,
            'next': self.get_next_link(),
            'previous': self.get_previous_link(),
            'last_page': total_pages,
            'results': data
        })


class ObjectsNumbersAPIViewSet(APIView):
    serializer_class = SocialMediaSerializer
    permission_classes = [AllowAny]
    def get(self, *args, **kwargs):
        try:
            filter = self.request.GET.get('filter', '')
            if filter == "sources":
                kb = KnowledgeBase.objects.filter(source__isnull=False)
                kbl = KnowledgeBaseLabelUser.objects.filter(knowledge_base__source__isnull=False)
            elif filter == "social_media":
                kb = KnowledgeBase.objects.filter(social_media__isnull=False)
                kbl = KnowledgeBaseLabelUser.objects.filter(knowledge_base__social_media__isnull=False)
            else:
                kb = KnowledgeBase.objects.all()
                kbl = KnowledgeBaseLabelUser.objects.all()

            all_count = kb.count()
            # real = kb.filter(category="real").count()
            # real = kbl.filter(label__name="حقیقت").count()

            only_haqiqat_kb_ids = KnowledgeBaseLabelUser.objects.values('knowledge_base') \
                .annotate(
                    total_labels=Count('id'),
                    haqiqat_labels=Count('id', filter=Q(label__name="حقیقت"))
                ).filter(total_labels=F('haqiqat_labels')).values_list('knowledge_base', flat=True)

            # و حالا گرفتن رکوردها
            real = kbl.filter(knowledge_base__in=only_haqiqat_kb_ids).count()
            dis = kbl.filter(label__name="فریب‌دهی").count()
            mis = kbl.filter(label__name="نادرست").count()
            mal = kbl.filter(label__name="مخرب").count()
            other =  kbl.exclude(Q(label__name="حقیقت") | Q(label__name="مخرب") | Q(label__name="فریب‌دهی") | Q(label__name="نادرست")).count()
            # other = max(0, all_count - (dis + real + mis + mal))


            def percent(count):
                return round((count / all_count) * 100, 2) if all_count > 0 else 0

            data = {
                "all": all_count,
                "real": real,
                "real_percent": percent(real),
                "dis": dis,
                "dis_percent": percent(dis),
                "mis": mis,
                "mis_percent": percent(mis),
                "mal": mal,
                "mal_percent": percent(mal),
                "other": other,
                "other_percent": percent(other),
            }

            return Response(data, status=status.HTTP_200_OK)
        except Exception as e:
            return Response({"error": "Something went wrong, try again - {}".format(e)}, status=status.HTTP_400_BAD_REQUEST)


class SocialmediaFullAPIViewSet(GenericAPIView):
    queryset = SocialMedia.objects.all()
    permission_classes = [AllowAny]
    pagination_class = CustomPagination
    serializer_class = SocialMediaSerializer
    filter_backends = [DjangoFilterBackend, SearchFilter, OrderingFilter]
    filterset_fields = ['title', 'created_at']
    search_fields = ['title', 'description']
    ordering_fields = ['id', 'created_at']
    def get(self, *args, **kwargs):
        sm = self.filter_queryset(SocialMedia.objects.all())
        serializer = self.serializer_class(sm, many=True, context={'request': self.request})
        filtered_data = [item for item in serializer.data if item['knowledge_base_items']]
        
        # Manual pagination
        page_size = self.pagination_class.page_size
        page_number = self.request.query_params.get('page', 1)
        try:
            page_number = int(page_number)
        except (ValueError, TypeError):
            page_number = 1
            
        start = (page_number - 1) * page_size
        end = start + page_size
        
        paginated_data = filtered_data[start:end]
        total_count = len(filtered_data)
        
        response_data = {
            'count': total_count,
            'next': f'?page={page_number + 1}' if end < total_count else None,
            'previous': f'?page={page_number - 1}' if page_number > 1 else None,
            'results': paginated_data
        }
        
        return Response(response_data, status=status.HTTP_200_OK)
    
    def post(self, *args, **kwargs):
        serializer = self.serializer_class(data=self.request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_200_OK)
        return Response(serializer.errors, status=status.HTTP_406_NOT_ACCEPTABLE)



class SocialmediaItemViewSet(APIView):
    serializer_class = SocialMediaSerializer
    permission_classes = [AllowAny]
    def get(self, *args, **kwargs):
        try:
            sm = SocialMedia.objects.get(id=self.kwargs["id"])
            serializer = self.serializer_class(sm,context={'request': self.request})
            return Response(serializer.data, status=status.HTTP_200_OK)
        except:
            return Response("SocialMedia not found or something went wrong, try again", status=status.HTTP_400_BAD_REQUEST)

    def put(self, *args, **kwargs):
        try:
            sm = SocialMedia.objects.get(id=self.kwargs["id"])
            serializer = self.serializer_class(sm, data=self.request.data, partial=True)
            if serializer.is_valid():
                obj = serializer.save()
                return Response(self.serializer_class(obj).data, status=status.HTTP_200_OK)
            return Response(serializer.errors, status=status.HTTP_406_NOT_ACCEPTABLE)
        except:
            return Response("SocialMedia not found or something went wrong, try again", status=status.HTTP_400_BAD_REQUEST)

    def delete(self, *args, **kwargs):
        try:
            sm = SocialMedia.objects.get(id=self.kwargs["id"])
            sm.delete()
            return Response("SocialMedia deleted.", status=status.HTTP_200_OK)
        except:
            return Response("SocialMedia not found or something went wrong, try again", status=status.HTTP_400_BAD_REQUEST)




class SourceFullAPIViewSet(GenericAPIView):
    queryset = Source.objects.all()
    permission_classes = [AllowAny]
    pagination_class = CustomPagination
    serializer_class = SourceFullSerializer
    filter_backends = [DjangoFilterBackend, SearchFilter, OrderingFilter]
    filterset_fields = ['title', 'created_at','category']
    search_fields = ['title', 'description']
    ordering_fields = ['id', 'created_at']

    def get_queryset(self):
        # Prefetch knowledge base items using the new related_name
        return Source.objects.prefetch_related(
            'knowledge_bases'
        ).annotate(
            kb_count=models.Count('knowledge_bases')
        ).filter(kb_count__gt=0).order_by('-id')  # Add default ordering by id

    def get(self, *args, **kwargs):
        queryset = self.get_queryset()
        filtered_queryset = self.filter_queryset(queryset)
        page = self.paginate_queryset(filtered_queryset)
        if page is not None:
            serializer = self.serializer_class(page, many=True, context={'request': self.request})
            return self.get_paginated_response(serializer.data)
        serializer = self.serializer_class(filtered_queryset, many=True, context={'request': self.request})
        return Response(serializer.data, status=status.HTTP_200_OK)



class SourceViewSet(GenericAPIView):
    serializer_class = SourceFullSerializer
    permission_classes = [AllowAny]
    pagination_class = CustomPagination
    def get(self, *args, **kwargs):
    
        sources = Source.objects.all()
        page = self.paginate_queryset(sources)
        if page is not None:
            serializer = self.serializer_class(page, many=True ,context={'request': self.request})
            return self.get_paginated_response(serializer.data)
        serializer = self.filter_queryset(Source.objects.all())
        return Response(serializer.data, status=status.HTTP_200_OK)
    
    def post(self, *args, **kwargs):
        serializer = CreateSourceSerializer(data=self.request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_200_OK)
        return Response(serializer.errors, status=status.HTTP_406_NOT_ACCEPTABLE)

class KnowledgeBaseFilter(django_filters.FilterSet):
    label_name = django_filters.CharFilter(method='filter_by_label_name')
    social_media__isnull = django_filters.BooleanFilter(field_name='social_media', lookup_expr='isnull')
    source__isnull = django_filters.BooleanFilter(field_name='source', lookup_expr='isnull')

        # فیلتر بین دو تاریخ
    date_time_pub__gte = django_filters.DateTimeFilter(field_name="date_time_pub", lookup_expr='gte')
    date_time_pub__lte = django_filters.DateTimeFilter(field_name="date_time_pub", lookup_expr='lte')

    is_news = django_filters.BooleanFilter(field_name="is_news")
    class Meta:
        model = KnowledgeBase
        fields = ['category', 'source', 'social_media', 'created_at', 'date_time_pub', 'is_news']

class SourceItemViewSet(GenericAPIView):
    queryset = Source.objects.all()
    permission_classes = [AllowAny]
    serializer_class = SourceSerializer
    filter_backends = [DjangoFilterBackend, SearchFilter, OrderingFilter]
    search_fields = ['title', 'body']
    ordering_fields = ['id', 'created_at']
    filterset_class = KnowledgeBaseFilter  # استفاده از فیلتر سفارشی
    def get_queryset(self):
        # کوئری‌ست پایه برای Source (برای اطمینان از وجود Source)
        return self.queryset
    def get_knowledge_base_queryset(self):
        # کوئری‌ست برای KnowledgeBase‌های مرتبط با Source
        source_id = self.kwargs.get("id")
        return KnowledgeBase.objects.filter(source_id=source_id)
    def get(self, *args, **kwargs):
        try:
            source = self.get_queryset().get(id=self.kwargs["id"])
            knowledge_base_queryset = self.filter_queryset(self.get_knowledge_base_queryset())
            serializer = SourceWithKBSerializer(
                source, context={'request': self.request,
                'filtered_knowledge_bases': knowledge_base_queryset
                })

            return Response(serializer.data, status=status.HTTP_200_OK)
        except Exception as e:
            return Response("Error - {}".format(e), status=status.HTTP_400_BAD_REQUEST)

    def put(self, *args, **kwargs):
        source = get_object_or_404(Source, id=self.kwargs["id"])

        serializer = EditSourceSerializer(source, data=self.request.data, partial=True)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_406_NOT_ACCEPTABLE)
        with transaction.atomic():
            obj = serializer.save()
            default_label_id = self.request.data.get("default_label")
            
            if default_label_id:
                try:
                    default_label = Label.objects.get(id=default_label_id)
                    knowledge_bases = KnowledgeBase.objects.filter(source=source)
                    KnowledgeBaseLabelUser.objects.filter(
                        knowledge_base__in=knowledge_bases, 
                        user__username= "system_user"
                    ).update(label=default_label)
                except Label.DoesNotExist:
                    return Response(
                        {"error": "Label with provided default_label ID not found"},
                        status=status.HTTP_400_BAD_REQUEST
                    )

        return Response(self.serializer_class(obj).data, status=status.HTTP_200_OK)

    def delete(self, *args, **kwargs):
        try:
            source = Source.objects.get(id=self.kwargs["id"])
            source.delete()
            return Response("Source deleted.", status=status.HTTP_200_OK)
        except:
            return Response("Source not found or something went wrong, try again", status=status.HTTP_400_BAD_REQUEST)

class GenerateSourceFilesView(APIView):
    def post(self, request):
        sources = Source.objects.filter(Q(file__isnull=True) | Q(file=''))
        for source in sources:
            kb_bodies = list(
                source.knowledge_bases.filter(body__isnull=False)
                .values_list('body', flat=True)
            )
            if not kb_bodies:
                continue

            data = [{"body": body} for body in kb_bodies]

            # تبدیل به JSON
            json_data = json.dumps(data, ensure_ascii=False, indent=2)
            file_name = f"source_file/source_{source.id}.json"

            # تبدیل string به فایل برای FileField
            content_file = ContentFile(json_data.encode('utf-8'), name=file_name)

            # آپلود در فایل‌فیلد
            source.file.save(file_name, content_file)
            source.save()
        return Response({"message": "Files generated for sources without file."}, status=status.HTTP_200_OK)


class LabelViewSet(APIView):
    serializer_class = LabelSerializer
    permission_classes = [AllowAny]
    def get(self, *args, **kwargs):
        labels = Label.objects.all()
        serializer = self.serializer_class(labels,many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)
    def post(self, *args, **kwargs):
        serializer = self.serializer_class(data=self.request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_200_OK)
        return Response(serializer.errors, status=status.HTTP_406_NOT_ACCEPTABLE)



class LabelItemViewSet(APIView):
    serializer_class = LabelSerializer
    permission_classes = [AllowAny]
    def get(self, *args, **kwargs):
        try:
            label = Label.objects.get(id=self.kwargs["id"])
            serializer = self.serializer_class(label)
            return Response(serializer.data, status=status.HTTP_200_OK)
        except:
            return Response("Label not found or something went wrong, try again", status=status.HTTP_400_BAD_REQUEST)

    def put(self, *args, **kwargs):
        try:
            label = Label.objects.get(id=self.kwargs["id"])
            serializer = self.serializer_class(label, data=self.request.data, partial=True)
            if serializer.is_valid():
                serializer.save()
                return Response(serializer.data, status=status.HTTP_200_OK)
            return Response(serializer.errors, status=status.HTTP_406_NOT_ACCEPTABLE)
        except:
            return Response("label not found or something went wrong, try again", status=status.HTTP_400_BAD_REQUEST)

    def delete(self, *args, **kwargs):
        try:
            label = Label.objects.get(id=self.kwargs["id"])
            label.delete()
            return Response("Label deleted.", status=status.HTTP_200_OK)
        except:
            return Response("Label not found or something went wrong, try again", status=status.HTTP_400_BAD_REQUEST)



class AddLabelViewSet(APIView):
    serializer_class = KnowledgeBaseLabelUserSerializer
    permission_classes = [IsAuthenticated]
    def delete(self, *args, **kwargs):
        data = self.request.data
        user = self.request.user
        label_id = data["label"]
        knowledge_base_id = data["knowledge_base"]
        try:
            instance = KnowledgeBaseLabelUser.objects.get(
                label_id=label_id,
                knowledge_base_id=knowledge_base_id,
                user=user
            )
            instance.delete()
            return Response({'message': 'Association removed successfully.'}, status=status.HTTP_204_NO_CONTENT)
        except KnowledgeBaseLabelUser.DoesNotExist:
            return Response({'error': 'Association not found.'}, status=status.HTTP_404_NOT_FOUND)


    def post(self, *args, **kwargs):
        data=self.request.data
        data["user"]=self.request.user.id
        user_type_names = self.request.user.user_type.values_list('name', flat=True)
        if not any(name in ['researcher', 'manager'] for name in user_type_names):
            return Response("You do not have permission to edit the element", status=status.HTTP_406_NOT_ACCEPTABLE)
        try:
            # Check if an object already exists for this knowledge_base and user
            kb_label_user = KnowledgeBaseLabelUser.objects.get(
                knowledge_base_id=data["knowledge_base"],
                user_id=data["user"]
            )
            # If exists, update the label
            kb_label_user.label_id = data["label"]
            kb_label_user.save()
            serializer = CreateKnowledgeBaseLabelUserSerializer(kb_label_user)
            return Response(serializer.data, status=status.HTTP_200_OK)

        except KnowledgeBaseLabelUser.DoesNotExist:
            # If it does not exist, create a new one
            serializer = CreateKnowledgeBaseLabelUserSerializer(data=data)
            if serializer.is_valid():
                serializer.save()
                return Response(serializer.data, status=status.HTTP_200_OK)
            return Response(serializer.errors, status=status.HTTP_406_NOT_ACCEPTABLE)



class AddSourceLabelViewSet(APIView):
    serializer_class = KnowledgeBaseLabelUserSerializer
    permission_classes = [IsAuthenticated]
    def delete(self, *args, **kwargs):
        data = self.request.data
        user = self.request.user
        label_id = data["label"]
        knowledge_base_id = data["knowledge_base"]
        try:
            instance = KnowledgeBaseLabelUser.objects.get(
                label_id=label_id,
                knowledge_base_id=knowledge_base_id,
                user=user
            )
            instance.delete()
            return Response({'message': 'Association removed successfully.'}, status=status.HTTP_204_NO_CONTENT)
        except KnowledgeBaseLabelUser.DoesNotExist:
            return Response({'error': 'Association not found.'}, status=status.HTTP_404_NOT_FOUND)


    def post(self, *args, **kwargs):
        data=self.request.data
        data["user"]=self.request.user.id
        serializer = CreateKnowledgeBaseLabelUserSerializer(data=data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_200_OK)
        return Response(serializer.errors, status=status.HTTP_406_NOT_ACCEPTABLE)


class KnowledgeBaseFilter(django_filters.FilterSet):
    label_name = django_filters.CharFilter(method='filter_by_label_name')
    social_media__isnull = django_filters.BooleanFilter(field_name='social_media', lookup_expr='isnull')
    source__isnull = django_filters.BooleanFilter(field_name='source', lookup_expr='isnull')

        # فیلتر بین دو تاریخ
    date_time_pub__gte = django_filters.DateTimeFilter(field_name="date_time_pub", lookup_expr='gte')
    date_time_pub__lte = django_filters.DateTimeFilter(field_name="date_time_pub", lookup_expr='lte')

    is_news = django_filters.BooleanFilter(field_name="is_news")
    class Meta:
        model = KnowledgeBase
        fields = ['category', 'source', 'social_media', 'created_at', 'label_name', 'date_time_pub', 'is_news']

    def filter_by_label_name(self, queryset, name, value):
        if value == 'برچسب دلخواه':
            return queryset.filter(
            knowledgebaselabeluser__label__name__in=[
                name for name in Label.objects.values_list('name', flat=True)
                if name not in ['حقیقت', 'مخرب', 'نادرست', 'فریب‌دهی']
            ]
        )
        else:
            # رفتار پیش‌فرض برای سایر label_nameها
            return queryset.filter(
                knowledgebaselabeluser__label__name=value
            )

class KnowledgeBaseFullAPIViewSet(GenericAPIView):
    queryset = KnowledgeBase.objects.all()
    permission_classes = [AllowAny]
    pagination_class = CustomPagination
    serializer_class = KnowledgeBaseSerializer
    filter_backends = [DjangoFilterBackend, SearchFilter, OrderingFilter]
    # filterset_fields = ['category', 'source', 'social_media', 'created_at']
    filterset_class = KnowledgeBaseFilter  # استفاده از فیلتر سفارشی
    search_fields = ['title', 'body']
    ordering_fields = ['id', 'created_at']
    

    def get(self, *args, **kwargs):
        print("MMMMMMMMMMMMMMMMM")
        kb = self.filter_queryset(KnowledgeBase.objects.all())
        page = self.paginate_queryset(kb)
        if page is not None:
            serializer = self.serializer_class(page, many=True ,context={'request': self.request})
            return self.get_paginated_response(serializer.data)
        serializer = self.filter_queryset(KnowledgeBase.objects.all())
        return Response(serializer.data, status=status.HTTP_200_OK)




class KnowledgeBaseViewSet(APIView):
    serializer_class = KnowledgeBaseSerializer
    permission_classes = [AllowAny]
    def get(self, *args, **kwargs):
        kb = KnowledgeBase.objects.all()
        serializer = self.serializer_class(kb,many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)
    def post(self, *args, **kwargs):
        serializer = AddKnowledgeBaseSerializer(data=self.request.data)

        if serializer.is_valid():
            item = serializer.save()

            # url = 'http://89.42.199.251:5682/text/kb/add_news'
            url = 'http://core-ai-sahaa:5682/text/kb/add_news'
            headers = {
                'sahaa-ai-api': 'WGhgR5dOAEc34MI0Zpi5C2Y3LyjwT9Ex',
                'Content-Type': 'application/json',
            }
            payload = {
                'category': item.category,
                'id': str(item.id),
                'body': item.body,
            }
            response = requests.post(url, params=payload, headers=headers)

            if response.status_code == 200:
                data = response.json()
                return Response(KnowledgeBaseSerializer(item).data, status=status.HTTP_201_CREATED)
            else:
                print("Status Code:", response.status_code)
                print("Response Text:", response.text)
                print("Request Payload:", payload)
                print("Request Headers:", headers)
                return Response("Failed to submit data!", status=status.HTTP_400_BAD_REQUEST)
        return Response(serializer.errors, status=status.HTTP_406_NOT_ACCEPTABLE)



class KnowledgeBaseItemViewSet(APIView):
    serializer_class = KnowledgeBaseSerializer
    permission_classes = [AllowAny]
    def get(self, *args, **kwargs):
        try:
            kb = KnowledgeBase.objects.get(id=self.kwargs["id"])
            serializer = self.serializer_class(kb,context={'request': self.request})
            return Response(serializer.data, status=status.HTTP_200_OK)
        except Exception as e:
            return Response("KnowledgeBase not found or something went wrong, try again - {}".format(e), status=status.HTTP_400_BAD_REQUEST)

    def put(self, *args, **kwargs):
        try:
            kb = KnowledgeBase.objects.get(id=self.kwargs["id"])
            serializer = AddKnowledgeBaseSerializer(kb, data=self.request.data, partial=True)
            if serializer.is_valid():
                item = serializer.save()

                # url = 'http://89.42.199.251:5682/text/kb/update_news'
                url = 'http://core-ai-sahaa:5682/text/kb/update_news'
                headers = {
                    'sahaa-ai-api': 'WGhgR5dOAEc34MI0Zpi5C2Y3LyjwT9Ex',
                    'Content-Type': 'application/json',
                }
                payload = {
                    'old_category': item.old_category,
                    'id': str(item.id),
                    'body': item.body,
                    'new_category': item.category,
                }

                response = requests.put(url, params=payload, headers=headers)

                if response.status_code == 200:
                    data = response.json()
                    print(data)
                    return Response(KnowledgeBaseSerializer(item).data)
                else:
                    print("Status Code:", response.status_code)
                    print("Response Text:", response.text)
                    print("Request Payload:", payload)
                    print("Request Headers:", headers)
                    return Response("Failed to submit data!", status=status.HTTP_400_BAD_REQUEST)
            return Response(serializer.errors, status=status.HTTP_406_NOT_ACCEPTABLE)
        except:
            return Response("KnowledgeBase not found or something went wrong, try again", status=status.HTTP_400_BAD_REQUEST)

    def delete(self, *args, **kwargs):

        try:
            kb_id = self.kwargs["id"]
            item = KnowledgeBase.objects.get(id=kb_id)
            # url = 'http://89.42.199.251:5682/text/kb/remove_news'
            url = 'http://core-ai-sahaa:5682/text/kb/remove_news'
            headers = {
                'sahaa-ai-api': 'WGhgR5dOAEc34MI0Zpi5C2Y3LyjwT9Ex',
                'Content-Type': 'application/json',
            }
            payload = {
                'category': item.category,
                'id': str(item.id),
            }
            response = requests.delete(url, params=payload, headers=headers)

            if response.status_code == 200:
                data = response.json()
                print(data)
                item.delete()
                return Response("KnowledgeBase deleted.", status=status.HTTP_200_OK)
            else:
                print("Status Code:", response.status_code)
                print("Response Text:", response.text)
                print("Request Payload:", payload)
                print("Request Headers:", headers)
                return Response("Failed to submit data!", status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            return Response(f"Unexpected error for kb_id={kb_id}: {e}", status=status.HTTP_400_BAD_REQUEST)




class SearchByID(APIView):
    permission_classes = [AllowAny]
    def get(self, *args, **kwargs):

        search_id = self.kwargs.get("id")
        batch_id = self.kwargs.get("batch_id")

        if search_id:
            search = SearchData.objects.filter(id=search_id).first()
        elif batch_id:
            search = SearchData.objects.filter(import_batch_id=batch_id).first()
        else:
            return Response({'error': 'Neither id nor batch_id provided'}, status=status.HTTP_400_BAD_REQUEST)

        simmilar_news_ids = search.ai_answer.get("ai_result", {}).get("simmilar_news") or []
        simmilar_analysis_ids = search.ai_answer.get("ai_result", {}).get("simmilar_analysis") or []

        simmilar_news = []
        simmilar_analysis = []
        if simmilar_news_ids:
            simmilar_news = KnowledgeBase.objects.filter(id__in=simmilar_news_ids)

        if simmilar_analysis_ids:
            simmilar_analysis = KnowledgeBase.objects.filter(id__in=simmilar_analysis_ids)

        return Response({
            "search": SearchSerializer(search).data,
            "simmilar_news": KnowledgeBaseSerializer(simmilar_news, many=True).data,
            "simmilar_analysis": KnowledgeBaseSerializer(simmilar_analysis, many=True).data,
        }, status=status.HTTP_200_OK)

# import logging
# import random
# import unicodedata
# import requests
# from requests.adapters import HTTPAdapter
# from urllib3.util.retry import Retry
# from django.db import connection
# import sys
# handler = logging.StreamHandler(sys.stdout)
# formatter = logging.Formatter("[%(asctime)s] %(levelname)s %(name)s: %(message)s")
# handler.setFormatter(formatter)
# logger.addHandler(handler)
# logger.setLevel(logging.DEBUG)
# logger = logging.getLogger(__name__)

class Search(APIView):
    permission_classes = [AllowAny]

    def post(self, request, *args, **kwargs):
        try:

            search_text = self.request.data.get('search')
            if not search_text:
                return Response({'error': 'Search field is required.'}, status=status.HTTP_400_BAD_REQUEST)

            # Prepare search data
            search_data = {'user': self.request.user.id if self.request.user.is_authenticated else None, 'text': search_text}
            serializer = SearchSerializer(data=search_data)
            if serializer.is_valid():
                search_item = serializer.save()
                logger.debug(f"[SEARCH] Search item saved with id={search_item.id}")
            else:
                return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

            # External AI API request
            # url = "http://89.42.199.251:5682/text/check_news"
            url = "http://core-ai-sahaa:5682/text/check_news"
            headers = {
                'sahaa-ai-api': 'WGhgR5dOAEc34MI0Zpi5C2Y3LyjwT9Ex',
                'Content-Type': 'application/json'
            }
            # headers = {
            #     'sahaa-ai-api': 'WGhgR5dOAEc34MI0Zpi5C2Y3LyjwT9Ex',
            #     'Content-Type': 'application/json',
            #     'Accept': 'application/json',
            #     'Accept-Encoding': 'gzip, deflate',
            #     'Accept-Language': 'en-US,en;q=0.9',
            #     'Origin': 'http://62.60.198.225:5682',
            #     'Referer': 'http://62.60.198.225:5682/docs',
            #     'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/139.0.0.0 Safari/537.36'
            # }

            payload = {'input_news': search_text}

            logger.info(f"[AI_REQUEST] Sending request to {url} | payload={payload}")
            start = time.time()

            response = requests.post(url, params=payload, headers=headers, timeout=(3, 60), proxies={"http": None, "https": None})
            elapsed = time.time() - start
            logger.info(f"[AI_REQUEST] Got response {response.status_code} in {elapsed:.2f}s")

            if response.status_code == 200:
                ai_result = response.json()
                try:
                    fact = KnowledgeBase.objects.get(id=ai_result['fact_id'])
                    fact_data = KnowledgeBaseSerializer(fact, context={'request': request}).data
                    print("fffffffffff", ai_result['simmilar_news'])
                    similar_kbs = KnowledgeBase.objects.filter(id__in=ai_result['simmilar_news']).select_related('source', 'social_media')
                    # شمارش
                    counts = {
                        "foreign_social": 0,
                        "foreign_sites": 0,
                        "internal_social": 0,
                        "internal_sites": 0
                    }

                    for kb in similar_kbs:
                        # شبکه اجتماعی
                        if kb.social_media:
                            if getattr(kb.social_media, "origin_type", "") == "foreign":
                                counts["foreign_social"] += 1
                            elif getattr(kb.social_media, "origin_type", "") == "domestic":
                                counts["internal_social"] += 1

                        # سایت خبرگذاری
                        if kb.source:
                            if getattr(kb.source, "origin_type", "") == "foreign":
                                counts["foreign_sites"] += 1
                            elif getattr(kb.source, "origin_type", "") == "domestic":
                                counts["internal_sites"] += 1
                    
                    # محاسبه درصد
                    total = len(similar_kbs) or 1
                    radar_chart = {k: round(v / total * 100, 2) for k, v in counts.items()}

                    lbls = []
                    for lbl in fact_data["labels"]:
                        lbl_item = {
                            "name": lbl["label_name"],
                            "count": lbl["count"],
                            "percentage": round((lbl["count"] / len(fact_data["labels"])) * 100, 2)
                        }
                        lbls.append(lbl_item)

                    chart_data = {
                        "pie_chart": lbls,
                        "radar_chart": radar_chart
                    }

                except:
                    fact_data = None

                    chart_data = {
                        "pie_chart": None,
                        "radar_chart": None
                    }



                combined_result = {
                    'id': str(search_item.id),
                    'search_text': search_text,
                    'ai_result': ai_result,
                    'fact_data': fact_data,
                    'chart_data': chart_data
                }
                if ai_result['result'] == "real":
                    search_item.result = "real"
                else:
                    search_item.result = "fake"
                search_item.processed = True
                search_item.ai_answer = combined_result
                search_item.save()
                return Response(combined_result, status=status.HTTP_200_OK)
            else:
                logger.warning(f"AI service failed | Status: {response.status_code} | Response: {response.text}")
                return Response({'error': 'Failed to fetch AI response.'}, status=status.HTTP_400_BAD_REQUEST)

        except Exception as e:
            print("EXCEPTION >>>", e)
            return Response({
                'error': 'Something went wrong, please try again.{}'.format(e)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR)


# class Search(APIView):
#     permission_classes = [AllowAny]
#     def post(self, request, *args, **kwargs):
#         try:
#             search_text = request.data.get('search')
#             if not search_text:
#                 return Response({'error': 'Search field is required.'}, status=status.HTTP_400_BAD_REQUEST)

#             # Prepare search data
#             search_data = {'user': self.request.user.id if self.request.user.is_authenticated else None, 'text': search_text}
#             serializer = SearchSerializer(data=search_data)
#             if serializer.is_valid():
#                 search_item = serializer.save()
#             else:
#                 return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

#             from requests.adapters import HTTPAdapter
#             from urllib3.util.retry import Retry

#             def requests_retry_session(
#                 retries=3,
#                 backoff_factor=1,
#                 status_forcelist=(500, 502, 503, 504),
#                 session=None,
#             ):
#                 session = session or requests.Session()
#                 retry = Retry(
#                     total=retries,
#                     read=retries,
#                     connect=retries,
#                     backoff_factor=backoff_factor,
#                     status_forcelist=status_forcelist,
#                     allowed_methods=frozenset(["GET", "POST"]),  # خیلی مهم: POST هم retry بشه
#                     raise_on_status=False
#                 )
#                 adapter = HTTPAdapter(max_retries=retry)
#                 session.mount("http://", adapter)
#                 session.mount("https://", adapter)
#                 return session
#             # External AI API request
#             url = 'http://89.42.199.251:5682/text/check_news'
#             headers = {
#                 'sahaa-ai-api': 'WGhgR5dOAEc34MI0Zpi5C2Y3LyjwT9Ex',
#                 'Content-Type': 'application/json',
#             }
#             payload = {'input_news': search_text}

#             session = requests_retry_session()

#             response = session.post(url, params=payload, headers=headers, timeout=(3, 90))
#             print("response", response.status_code, response.text)
#             if response.status_code == 200:
#                 ai_result = response.json()

#                 try:
#                     fact = KnowledgeBase.objects.get(id=ai_result['fact_id'])
#                     fact_data = KnowledgeBaseSerializer(fact, context={'request': request}).data

#                     if_foreign = True
#                     for char in fact_data["source"]["title"]:
#                         if char.isalpha():
#                             try:
#                                 name = unicodedata.name(char)
#                                 if 'LATIN' not in name:
#                                     if_foreign = False
#                                     break  # No need to continue if we already know
#                             except ValueError:
#                                 if_foreign = False
#                                 break
#                     if if_foreign:
#                         foreign_social = 80
#                         foreign_sites = 95
#                         internal_social = 20
#                         internal_sites = 35
#                     else:
#                         foreign_social = 20
#                         foreign_sites = 35
#                         internal_social = 80
#                         internal_sites = 95

#                     def update_with_tolerance(value, tolerance=5):
#                         change = random.randint(0, tolerance)
#                         return value + change if random.choice([True, False]) else value - change

#                     radar_chart = {
#                         "foreign_social": update_with_tolerance(foreign_social),
#                         "foreign_sites": update_with_tolerance(foreign_sites),
#                         "internal_social": update_with_tolerance(internal_social),
#                         "internal_sites": update_with_tolerance(internal_sites)
#                     }

#                     lbls = []
#                     for lbl in fact_data["labels"]:
#                         lbl_item = {
#                             "name": lbl["label_name"],
#                             "count": lbl["count"],
#                             "percentage": round((lbl["count"] / len(fact_data["labels"])) * 100, 2)
#                         }
#                         lbls.append(lbl_item)

#                     chart_data = {
#                         "pie_chart": lbls,
#                         "radar_chart": radar_chart
#                     }

#                 except:
#                     fact_data = None

#                     chart_data = {
#                         "pie_chart": None,
#                         "radar_chart": None
#                     }



#                 combined_result = {
#                     'id': str(search_item.id),
#                     'search_text': search_text,
#                     'ai_result': ai_result,
#                     'fact_data': fact_data,
#                     'chart_data': chart_data
#                 }
#                 if ai_result['result'] == "real":
#                     search_item.result = "real"
#                 else:
#                     search_item.result = "fake"
#                 search_item.processed = True
#                 search_item.ai_answer = combined_result
#                 search_item.save()
#                 return Response(combined_result, status=status.HTTP_200_OK)
#             else:
#                 logger.warning(f"AI service failed | Status: {response.status_code} | Response: {response.text}")
#                 return Response({'error': 'Failed to fetch AI response.'}, status=status.HTTP_400_BAD_REQUEST)

#         except Exception as e:
#             logger.exception("Unexpected error occurred during search")
#             return Response({'error': 'Something went wrong, please try again.{}'.format(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    def get(self, *args, **kwargs):
        try:
            search = SearchData.objects.filter(user=self.request.user)
            serializer = SearchSerializer(search,many=True)
            return Response(serializer.data, status=status.HTTP_200_OK)
        except Exception as e:
            return Response("Error - {}".format(e), status=status.HTTP_400_BAD_REQUEST)





class UploadSearch(APIView):
    permission_classes = [AllowAny]
    def post(self, request, *args, **kwargs):
        source_data = self.request.data
        socialmedia_serializer = CreateSocialMedia(data=source_data)
        if socialmedia_serializer.is_valid():
            socialmedia_item = socialmedia_serializer.save()
            done_item = 0
            done_item_in_server = 0

            try:
                json_file = self.request.FILES.get('file')
                print(json_file.read())
                json_file.seek(0)
                try:
                    file_data = json.load(json_file)
                    for obj in file_data:
                        search_text = obj["body"]

                        # Prepare search data
                        search_data = {'user': self.request.user.id if self.request.user.is_authenticated else None,
                                       'text': search_text}
                        serializer = SearchSerializer(data=search_data)
                        if serializer.is_valid():
                            search_item = serializer.save()
                            done_item += 1

                        # External AI API request
                        # url = 'http://89.42.199.251:5682/text/check_news'
                        url = 'http://core-ai-sahaa:5682/text/check_news'
                        headers = {
                            'sahaa-ai-api': 'WGhgR5dOAEc34MI0Zpi5C2Y3LyjwT9Ex',
                            'Content-Type': 'application/json',
                        }
                        payload = {'input_news': search_text}

                        response = requests.post(url, params=payload, headers=headers)

                        if response.status_code == 200:
                            ai_result = response.json()
                            done_item_in_server += 1

                            try:
                                fact = KnowledgeBase.objects.get(id=ai_result['fact_id'])
                                fact_data = KnowledgeBaseSerializer(fact).data
                            except:
                                fact_data = None
                            combined_result = {
                                'id': str(search_item.id),
                                'search_text': search_text,
                                'ai_result': ai_result,
                                'fact_data': fact_data
                            }
                            if ai_result['result'] == "real":
                                search_item.result = "real"
                            else:
                                search_item.result = "fake"
                            search_item.processed = True
                            search_item.ai_answer = combined_result
                            search_item.save()

                    final_data = {
                        "social_media": socialmedia_serializer.data,
                        "backend_kb_added": done_item,
                        "AI_kb_added": done_item_in_server
                    }
                    return Response(final_data, status=status.HTTP_200_OK)


                except json.JSONDecodeError:
                    return Response({"error": "Invalid JSON file"}, status=status.HTTP_400_BAD_REQUEST)

            except Exception as e:
                logger.exception("Unexpected error occurred during search")
                return Response({'error': 'Something went wrong, please try again.{}'.format(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

        return Response(socialmedia_serializer.errors, status=status.HTTP_406_NOT_ACCEPTABLE)


logger = logging.getLogger(__name__)
class UploadSourceFile(APIView):
    permission_classes = [AllowAny]
    def post(self, request, *args, **kwargs):
        try:
            # --- اعتبارسنجی منبع
            source_data = request.data
            source_title = source_data.get("title")
            if not source_title:
                return Response(
                    {"error": "Title is required"},
                    status=status.HTTP_400_BAD_REQUEST
                )
            source_item, created = Source.objects.get_or_create(
                title=source_title,

                defaults={
                    "description": source_data.get("description"),
                    "category": source_data.get("category"),
                    "photo": source_data.get("photo"),
                    "file": source_data.get("file"),
                    "source_uri": source_data.get("source_uri"),
                    "source_data_type": source_data.get("source_data_type"),
                }
            )
         

            # --- گرفتن فایل
            json_file = request.FILES.get("file")
            if not json_file:
                return Response(
                    {"error": "No JSON file provided."},
                    status=status.HTTP_400_BAD_REQUEST
                )

            try:
                file_data = json.load(json_file)
                if not isinstance(file_data, list):
                    return Response(
                        {"error": "JSON file must contain a list of objects."},
                        status=status.HTTP_400_BAD_REQUEST
                    )
            except json.JSONDecodeError as e:
                return Response(
                    {"error": f"Invalid JSON file: {str(e)}"},
                    status=status.HTTP_400_BAD_REQUEST
                )

            # --- پردازش آبجکت‌ها
            created_items = []
            failed_items = []
            for idx, obj in enumerate(file_data, start=1):
                uri = obj.get("uri")
                kb_data = {
                    "title": obj.get("title"),
                    "body": obj.get("body"),
                    "url": obj.get("url"),
                    "source": source_item.id,
                    "category": "real",
                    "date_time_pub": obj.get("date_time_pub"),
                    "image": obj.get("image"),
                    "uri": uri,
                    "lang": obj.get("lang"),
                    "is_duplicate": obj.get("isDuplicate"),
                    "data_type": obj.get("dataType"),
                    "sim": obj.get("sim"),
                    "sentiment": obj.get("sentiment"),
                    "wgt": obj.get("wgt"),
                    "relevance": obj.get("relevance"),
                    "authors": obj.get("authors"),
                }
                if uri and KnowledgeBase.objects.filter(uri=uri).exists():
                    continue 
                serializer = AddKnowledgeBaseSerializer(data=kb_data)
                if serializer.is_valid():
                    created_items.append(serializer.save().id)
                else:
                    failed_items.append({"index": idx, "errors": serializer.errors})

            final_data = {
                "source": {
                    "id": source_item.id,
                    "title": source_item.title,
                    "created": created,  # True = سورس جدید ساخته شده، False = قبلاً وجود داشته
                },
                "created_count": len(created_items),
                "created_items": created_items,                                                       
                "failed_count": len(failed_items),
                "failed_items": failed_items,  # برای دیباگ کاربر
            }

            return Response(final_data, status=status.HTTP_201_CREATED)

        except Exception as e:
            logger.exception("Unexpected error occurred during upload")
            return Response(
                {"error": "Something went wrong on the server.", "details": str(e)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
        

class ImportKnowledgeBaseExcelView(APIView):
    parser_classes = [MultiPartParser]
    def clean_text(self, text: str) -> str:
        """حذف بک‌اسلش‌های اضافه و quote اضافی از متن"""
        if not text:
            return ""
        text = text.replace('\\"', '"').strip()
        if text.startswith('"') and text.endswith('"'):
            text = text[1:-1].strip()
        return text
    
    def post(self, request):
        excel_file = request.FILES.get("file")
        if not excel_file:
            return Response({"error": "Excel file is required."}, status=status.HTTP_400_BAD_REQUEST)

        try:
            wb = openpyxl.load_workbook(filename=excel_file, read_only=True)
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))
        except Exception as e:
            return Response({"error": f"Failed to read Excel file: {str(e)}"}, status=status.HTTP_400_BAD_REQUEST)

        if not rows or len(rows) < 2:
            return Response({"error": "Excel file is empty or has no data rows."}, status=status.HTTP_400_BAD_REQUEST)

        headers = [str(h).strip() if h else "" for h in rows[0]]
        data_rows = rows[1:]

        # پیدا کردن ایندکس فیلدها
        def get_idx(field_name):
            return headers.index(field_name) if field_name in headers else None

        idx_map = {
            "title": get_idx("title"),
            "body": get_idx("body"),
            "url": get_idx("url"),
            "date_time_pub": get_idx("date_time_pub"),
            "source_uri": get_idx("source_uri"),
            "source_title": get_idx("source_title"),
            "uri": get_idx("uri"),
            "category": get_idx("category"),
            "is_news": get_idx("is_news"),
            "sim": get_idx("sim"),
            "sentiment": get_idx("sentiment"),
            "wgt": get_idx("wgt"),
            "relevance": get_idx("relevance"),
            "authors": get_idx("authors"),
            "image": get_idx("image"),
            "lang": get_idx("lang"),
            "is_duplicate": get_idx("is_duplicate"),
            "data_type": get_idx("data_type"),
        }

        kb_objects = []
        created_count = 0
        skipped_count = 0

        with transaction.atomic():
            for row in data_rows:
                # فیلدهای اجباری
                title = row[idx_map["title"]] if idx_map["title"] is not None else None
                body = row[idx_map["body"]] if idx_map["body"] is not None else None
                
                title = self.clean_text(title)
                body = self.clean_text(body)

                if not title or not body:
                    skipped_count += 1
                    continue  # اگر title یا body موجود نباشه، رد کن

                # بررسی source
                source_obj = None
                source_uri = row[idx_map["source_uri"]] if idx_map["source_uri"] is not None else None
                source_title = row[idx_map["source_title"]] if idx_map["source_title"] is not None else None

                # اگر source_uri موجود باشه، اول روی اون چک می‌کنیم
                if source_uri:
                    source_obj = Source.objects.filter(source_uri=source_uri).first()
                    # اگه پیدا نشد و title هم هست، بسازیم
                    if not source_obj and source_title:
                        source_obj = Source.objects.create(title=source_title.strip(), source_uri=source_uri.strip())
                # اگر نه uri باشه ولی title باشه، می‌تونیم فقط title رو بسازیم یا اسکیپ کنیم
                elif source_title:
                    source_obj = Source.objects.filter(title__iexact=source_title.strip()).first()
                    if not source_obj:
                        source_obj = Source.objects.create(title=source_title.strip())

                # اگر نه uri و نه title بود، اسکیپ می‌کنیم این رکورد
                if not source_obj:
                    # می‌تونیم log هم کنیم که رکورد اسکیپ شد
                    continue

                # بررسی uri خبر برای جلوگیری از رکورد تکراری
                news_uri = row[idx_map["uri"]] if idx_map["uri"] is not None else None
                if news_uri and KnowledgeBase.objects.filter(uri=news_uri).exists():
                    skipped_count += 1
                    continue

                kb_data = {
                    "title": title,
                    "body": body,
                    "source": source_obj,
                    "url": row[idx_map["url"]] if idx_map["url"] is not None else None,
                    "date_time_pub": row[idx_map["date_time_pub"]] if idx_map["date_time_pub"] is not None else None,
                    "category": row[idx_map["category"]] if idx_map["category"] is not None else None,
                    "is_news": row[idx_map["is_news"]] if idx_map["is_news"] is not None else None,
                    "sim": row[idx_map["sim"]] if idx_map["sim"] is not None else None,
                    "sentiment": row[idx_map["sentiment"]] if idx_map["sentiment"] is not None else None,
                    "wgt": row[idx_map["wgt"]] if idx_map["wgt"] is not None else None,
                    "relevance": row[idx_map["relevance"]] if idx_map["relevance"] is not None else None,
                    "authors": row[idx_map["authors"]] if idx_map["authors"] is not None else None,
                    "image": row[idx_map["image"]] if idx_map["image"] is not None else None,
                    "lang": row[idx_map["lang"]] if idx_map["lang"] is not None else None,
                    "is_duplicate": row[idx_map["is_duplicate"]] if idx_map["is_duplicate"] is not None else None,
                    "data_type": row[idx_map["data_type"]] if idx_map["data_type"] is not None else None,
                    "uri": news_uri,
                }

                kb_objects.append(KnowledgeBase(**kb_data))
                created_count += 1

        # ذخیره bulk
        if kb_objects:
            saved_kbs = KnowledgeBase.objects.bulk_create(kb_objects)
            #we dont need it with autoLabeling
            # assign_default_labels_to_kbs(saved_kbs)
            create_kb_process_status(saved_kbs)

        return Response({
            "message": "File imported successfully.",
            "inserted_count": created_count,
            "skipped_count": skipped_count
        }, status=status.HTTP_201_CREATED)

# class UploadSourceFile(APIView):
#     permission_classes = [AllowAny]
#     def post(self, request, *args, **kwargs):
#         source_data = self.request.data
#         source_serializer = CreateSourceSerializer(data=source_data)
#         if source_serializer.is_valid():
#             source_item = source_serializer.save()
#             print("ssssssssssss", source_item)
#             done_item = 0
#             done_item_in_server = 0
#             try:
#                 json_file = self.request.FILES.get('file')
#                 print("jjjjjjjjjjjjjjjjj",json_file.read())
#                 json_file.seek(0)
#                 try:
#                     file_data = json.load(json_file)
#                     for obj in file_data:
#                         search_text = obj["body"]

#                         # Prepare search data
#                         kb_data = {
#                             'user': self.request.user.id if self.request.user.is_authenticated else None,
#                             'title': search_text,
#                             'body': search_text,
#                             'source': source_item.id,
#                             'category': "real"
#                         }
#                         serializer = AddKnowledgeBaseSerializer(data=kb_data)
#                         if serializer.is_valid():
#                             item = serializer.save()
#                             done_item += 1
#                             url = 'http://89.42.199.251:5682/text/kb/add_news'
#                             headers = {
#                                 'sahaa-ai-api': 'WGhgR5dOAEc34MI0Zpi5C2Y3LyjwT9Ex',
#                                 'Content-Type': 'application/json',
#                             }
#                             payload = {
#                                 'category': item.category,
#                                 'id': str(item.id),
#                                 'body': item.body,
#                             }
#                             response = requests.post(url, params=payload, headers=headers)
#                             if response.status_code == 200:
#                                 data = response.json()
#                                 done_item_in_server += 1

#                     final_data = {
#                         "source": source_serializer.data,
#                         "backend_kb_added": done_item,
#                         "AI_kb_added": done_item_in_server
#                     }
#                     return Response(final_data, status=status.HTTP_200_OK)

#                 except json.JSONDecodeError as e:
#                     return Response({"error": "Invalid JSON file - {}".format(e)}, status=status.HTTP_400_BAD_REQUEST)

#             except Exception as e:
#                 logger.exception("Unexpected error occurred during search")
#                 return Response({'error': 'Something went wrong, please try again.{}'.format(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

#         return Response(source_serializer.errors, status=status.HTTP_406_NOT_ACCEPTABLE)






#----------------------




class MediaSearch(APIView):
    permission_classes = [AllowAny]
    def post(self, *args, **kwargs):
        try:
            data = self.request.data
            search = data['search']
            # todo - save search
            # todo - connect to AI and get data
            # todo - response data

            response_data = {
                "id": 243,
                "keywords": {"اعتراض", "نظام ـ سلامت"},
                "locations": {"ایران", "تهران", "خوزستان"},
                "possibilities": [
                    {"probability": "حقیقت", "percentage": 17.2},
                    {"probability": "دروغ", "percentage": 16.5},
                    {"probability": "فریب", "percentage": 36.6},
                    {"probability": "سوءقصد", "percentage": 28.9}
                ],
                "websites": [
                    {"website": "وبسایت‌های داخلی", "percentage": 60.4},
                    {"website": "شبکه‌های اجتماعی خارجی", "percentage": 92.2},
                    {"website": "وبسایت‌های خارجی", "percentage": 100},
                    {"website": "شبکه‌های اجتماعی داخلی", "percentage": 34.1}
                ],
                "posts": [
                    {
                        "title": "تجمع اعتراضی پرستاران مقابل استانداری خوزستان",
                        "image": "https://encrypted-tbn0.gstatic.com/images?q=tbn:ANd9GcQoBujouyZLFZC5nyf7TIUPqqEkNqLyXeabew&s",
                        "text": "امروز (30 آبان)، جمعی از پرستاران خوزستان با تجمع مقابل استانداری خواستار رسیدگی به مطالبات قانونی خود شدند. این پرستاران اصلی‌ترین خواسته‌های خود را برقراری فوق العاده خاص با ضریب 3 و گنجاندن آن در حکم، اجرای دقیق تعرفه‌گذاری عادلانه، اجرای قانون مشاغل سخت و زیان‌آور، ممنوعیت اضافه‌کار اجباری و برقراری امنیت شغلی عنوان کردند.",
                        "date": "June 7, 2021"
                    },
                    {
                        "title": "تجمع اعتراضی پرستاران مقابل استانداری خوزستان",
                        "image": "https://encrypted-tbn0.gstatic.com/images?q=tbn:ANd9GcQoBujouyZLFZC5nyf7TIUPqqEkNqLyXeabew&s",
                        "text": "امروز (30 آبان)، جمعی از پرستاران خوزستان با تجمع مقابل استانداری خواستار رسیدگی به مطالبات قانونی خود شدند. این پرستاران اصلی‌ترین خواسته‌های خود را برقراری فوق العاده خاص با ضریب 3 و گنجاندن آن در حکم، اجرای دقیق تعرفه‌گذاری عادلانه، اجرای قانون مشاغل سخت و زیان‌آور، ممنوعیت اضافه‌کار اجباری و برقراری امنیت شغلی عنوان کردند.",
                        "date": "June 7, 2021"
                    }
                ],
                "created_data": "Feb 25, 2025"
            }

            return Response(response_data, status=status.HTTP_200_OK)
        except:
            return Response("Something went wrong please try again.", status=status.HTTP_400_BAD_REQUEST)






class Answer(APIView):
    permission_classes = [AllowAny]
    def get(self, *args, **kwargs):
        try:
            #search = Search.objects.get(id=self.kwargs["id"])
            # todo - connect to AI and get search object
            # todo - response data

            data = {
                "id": 243,
                "keywords": {"اعتراض","نظام ـ سلامت"},
                "locations": {"ایران","تهران","خوزستان"},
                "possibilities": [
                    {"probability": "حقیقت", "percentage": 17.2},
                    {"probability": "دروغ", "percentage": 16.5},
                    {"probability": "فریب", "percentage": 36.6},
                    {"probability": "سوءقصد", "percentage": 28.9}
                ],
                "websites": [
                    {"website": "وبسایت‌های داخلی", "percentage": 60.4},
                    {"website": "شبکه‌های اجتماعی خارجی", "percentage": 92.2},
                    {"website": "وبسایت‌های خارجی", "percentage": 100},
                    {"website": "شبکه‌های اجتماعی داخلی", "percentage": 34.1}
                ],
                "posts": [
                    {
                        "title": "تجمع اعتراضی پرستاران مقابل استانداری خوزستان",
                        "image": "https://encrypted-tbn0.gstatic.com/images?q=tbn:ANd9GcQoBujouyZLFZC5nyf7TIUPqqEkNqLyXeabew&s",
                        "text": "امروز (30 آبان)، جمعی از پرستاران خوزستان با تجمع مقابل استانداری خواستار رسیدگی به مطالبات قانونی خود شدند. این پرستاران اصلی‌ترین خواسته‌های خود را برقراری فوق العاده خاص با ضریب 3 و گنجاندن آن در حکم، اجرای دقیق تعرفه‌گذاری عادلانه، اجرای قانون مشاغل سخت و زیان‌آور، ممنوعیت اضافه‌کار اجباری و برقراری امنیت شغلی عنوان کردند.",
                        "date": "June 7, 2021"
                    },
                    {
                        "title": "تجمع اعتراضی پرستاران مقابل استانداری خوزستان",
                        "image": "https://encrypted-tbn0.gstatic.com/images?q=tbn:ANd9GcQoBujouyZLFZC5nyf7TIUPqqEkNqLyXeabew&s",
                        "text": "امروز (30 آبان)، جمعی از پرستاران خوزستان با تجمع مقابل استانداری خواستار رسیدگی به مطالبات قانونی خود شدند. این پرستاران اصلی‌ترین خواسته‌های خود را برقراری فوق العاده خاص با ضریب 3 و گنجاندن آن در حکم، اجرای دقیق تعرفه‌گذاری عادلانه، اجرای قانون مشاغل سخت و زیان‌آور، ممنوعیت اضافه‌کار اجباری و برقراری امنیت شغلی عنوان کردند.",
                        "date": "June 7, 2021"
                    }
                ],
                "created_data": "Feb 25, 2025"
            }

            return Response(data, status=status.HTTP_200_OK)
        except:
            return Response("Something went wrong please try again.", status=status.HTTP_400_BAD_REQUEST)





"""
    # Get all data
    data = SearchData.objects()
    # Create a data
    data = SearchData(text="nima", email="nima@example.com")
    data.save()
    
    
    get method example:
    
    def get(self, *args, **kwargs):
       data = SearchData.objects()  # Get all MongoDB documents
       serializer = SearchDataSerializer(data, many=True)
       return Response(serializer.data)
"""





""" 
response_data = {
    "id": 243,
    "keywords": {"اعتراض", "نظام ـ سلامت"},
    "locations": {"ایران", "تهران", "خوزستان"},
    "possibilities": [
        {"probability": "حقیقت", "percentage": 17.2},
        {"probability": "دروغ", "percentage": 16.5},
        {"probability": "فریب", "percentage": 36.6},
        {"probability": "سوءقصد", "percentage": 28.9}
    ],
    "websites": [
        {"website": "وبسایت‌های داخلی", "percentage": 60.4},
        {"website": "شبکه‌های اجتماعی خارجی", "percentage": 92.2},
        {"website": "وبسایت‌های خارجی", "percentage": 100},
        {"website": "شبکه‌های اجتماعی داخلی", "percentage": 34.1}
    ],
    "posts": [
        {
            "title": "تجمع اعتراضی پرستاران مقابل استانداری خوزستان",
            "image": "https://encrypted-tbn0.gstatic.com/images?q=tbn:ANd9GcQoBujouyZLFZC5nyf7TIUPqqEkNqLyXeabew&s",
            "text": "امروز (30 آبان)، جمعی از پرستاران خوزستان با تجمع مقابل استانداری خواستار رسیدگی به مطالبات قانونی خود شدند. این پرستاران اصلی‌ترین خواسته‌های خود را برقراری فوق العاده خاص با ضریب 3 و گنجاندن آن در حکم، اجرای دقیق تعرفه‌گذاری عادلانه، اجرای قانون مشاغل سخت و زیان‌آور، ممنوعیت اضافه‌کار اجباری و برقراری امنیت شغلی عنوان کردند.",
            "date": "June 7, 2021"
        },
        {
            "title": "تجمع اعتراضی پرستاران مقابل استانداری خوزستان",
            "image": "https://encrypted-tbn0.gstatic.com/images?q=tbn:ANd9GcQoBujouyZLFZC5nyf7TIUPqqEkNqLyXeabew&s",
            "text": "امروز (30 آبان)، جمعی از پرستاران خوزستان با تجمع مقابل استانداری خواستار رسیدگی به مطالبات قانونی خود شدند. این پرستاران اصلی‌ترین خواسته‌های خود را برقراری فوق العاده خاص با ضریب 3 و گنجاندن آن در حکم، اجرای دقیق تعرفه‌گذاری عادلانه، اجرای قانون مشاغل سخت و زیان‌آور، ممنوعیت اضافه‌کار اجباری و برقراری امنیت شغلی عنوان کردند.",
            "date": "June 7, 2021"
        }
    ],
    "created_data": "Feb 25, 2025"
}
"""


class DownloadSearchData(APIView):
    permission_classes = [IsAuthenticated]
    
    def get(self, request, *args, **kwargs):
        try:
            logger.info("Starting file download for user: %s", request.user.id)
            
            # Get user's search data
            search_data = SearchData.objects.filter(user=request.user).order_by('-created_at')
            logger.info("Found %d search records", search_data.count())
            
            # Create Excel file in memory
            output = BytesIO()
            workbook = xlsxwriter.Workbook(output)
            worksheet = workbook.add_worksheet('Search History')
            
            # Add headers with formatting
            headers = [
                'Search ID', 'Search Text', 'Result', 'Created At', 
                'Updated At', 'Processed', 'AI Result', 'Fact ID', 'Confidence'
            ]
            
            header_format = workbook.add_format({
                'bold': True,
                'bg_color': '#D9E1F2',
                'border': 1,
                'align': 'center',
                'valign': 'vcenter'
            })
            
            cell_format = workbook.add_format({
                'align': 'center',
                'valign': 'vcenter',
                'text_wrap': True
            })
            
            # Set column widths
            widths = [15, 50, 15, 20, 20, 10, 15, 15, 15]
            for col, width in enumerate(widths):
                worksheet.set_column(col, col, width)
                worksheet.write(0, col, headers[col], header_format)
            
            # Write data rows
            for row, item in enumerate(search_data, start=1):
                ai_result = item.ai_answer.get('ai_result', {}) if item.ai_answer else {}
                
                row_data = [
                    str(item.id),
                    item.text,
                    item.result or 'Not processed',
                    item.created_at.strftime('%Y-%m-%d %H:%M:%S'),
                    item.updated_at.strftime('%Y-%m-%d %H:%M:%S'),
                    'Yes' if item.processed else 'No',
                    ai_result.get('result', 'N/A'),
                    ai_result.get('fact_id', 'N/A'),
                    ai_result.get('confidence', 'N/A')
                ]
                
                for col, value in enumerate(row_data):
                    worksheet.write(row, col, value, cell_format)
            
            # Add autofilter
            worksheet.autofilter(0, 0, len(search_data), len(headers) - 1)
            
            # Freeze panes
            worksheet.freeze_panes(1, 0)
            
            workbook.close()
            
            # Create the response
            output.seek(0)
            response = HttpResponse(
                output.read(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = 'attachment; filename=search_history.xlsx'
            
            return response
            
        except Exception as e:
            logger.exception("Error in download: %s", str(e))
            return Response(
                {'error': 'Failed to generate Excel file: {}'.format(str(e))},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
class AssignDefaultTruthLabelView(APIView): 
    @transaction.atomic
    def get(self, request):
        # بررسی دسترسی کاربر
        user_type_names = request.user.user_type.values_list('name', flat=True)
        if not any(name in ['researcher', 'manager'] for name in user_type_names):
            return Response({"error": "شما اجازه انجام این عملیات را ندارید"}, status=status.HTTP_403_FORBIDDEN)

        try:
            # دریافت یا ایجاد کاربر پیش‌فرض
            system_user, created  = User.objects.get_or_create(email="system_user@yourapp.com", defaults={"username": "system_user", "name": "System User", "password": "system_user"})
            
            # دریافت یا ایجاد لیبل "حقیقت"
            truth_label, _ = Label.objects.get_or_create(name="حقیقت")
        except Exception as e:
            return Response({"error": f"خطا در ایجاد یا دریافت کاربر یا لیبل: {str(e)}"}, status=status.HTTP_400_BAD_REQUEST)
        
        kb_ids_with_label = KnowledgeBaseLabelUser.objects.filter(
                label=truth_label,
                user=system_user
            ).values_list('knowledge_base_id', flat=True)

        new_objects = []
        for kb in KnowledgeBase.objects.exclude(id__in=kb_ids_with_label):
            new_objects.append(
                KnowledgeBaseLabelUser(
                    knowledge_base=kb,
                    label=truth_label,
                    user=system_user
                )
            )

        KnowledgeBaseLabelUser.objects.bulk_create(new_objects, batch_size=2000)

        return Response({
            "status": "success",
            "message": f"{len(new_objects)} records assigned with label 'حقیقت' by system user.",
            "user_id": system_user.id
        })
    
class NewsAPIView(APIView):
    def get(self, request):
        # دریافت API Key
        api_key = os.getenv("NEWS_AI_KEY")
        if not api_key:
            return Response(
                {"error": "API KEY تنظیم نشده است"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
        
        url = "https://eventregistry.org/api/v1/article/getArticles"

        
        params = {
            "apiKey": api_key,
            "lang": "fas",
            # "conceptUri": "http://en.wikipedia.org/wiki/Iran",
            # "lang": "mul",
            "dateStart": "2025-09-01",
            "dateEnd": "2025-09-15",
            # "sentiment": "negative",
            # "locationUri": "http://en.wikipedia.org/wiki/Iran"
            
            "isDuplicateFilter": "skipDuplicates",
            "dataType": ["news"],
            # "dataType": ["news", "blog", "pr"],
            # "sourceUri": [ 
            #     "farsnews.ir", "tasnimnews.com", "isna.ir", "sharghdaily.com",
            #     "irna.ir", "mehrnews.com",
            #     "tabnak.ir", "mostaghelonline.com",
            #     "hamshahrionline.ir",

                
            #     # "nytimes.com",
            #     # "foxnews.com", "msnbc.com",
            #     # "telegraph.co.uk", "dailymail.co.uk", "bbc.com", "theguardian.com",
            #     # "israelhayom.com", "israelnationalnews.com", "haaretz.com",
            #     # "jpost.com", "sputniknews.com", "novayagazeta.ru", "globaltimes.cn",
            #     # "peopledaily.com.cn"
            #     # "reuters.com" 
            #     # "aljazeera.net", "aljazeera.com"

            # ],
            "sourceUri": [ 
                
                "radiofarda.com", "ir.voanews.com", "iranintl.com", "iranintl-website.vercel.app",
                "bbc.com/persian", 
            ],
            # "sourceUri": [ 
            #     "apnews.com",
            #     "upi.com",
            #     "edition.cnn.com", "cnn.com", 
            #     "bbc.co.uk",
            #     "france24.com",
            #     "alarabiya.net", "english.alarabiya.net",
            #     "almanar.com.lb",
            #     "bloomberg.com",
            #     "rt.com",
            #     "wsj.com",
            #     "washingtonpost.com"
            # ],
            "sortBy": "rel",
            "articlesPage": 1,
            "articlesCount": 100,
        }

        all_articles = []
        while True:
            print(f"🔎 Fetching page {params['articlesPage']}...")

            try:
                response = requests.get(url, params=params)
                
                if response.status_code == 200:
                    data = response.json()
                    # print("ffffffffffffffffffff", data)
                    # استخراج نتایج و اطلاعات صفحه
                    articles = data.get("articles", {}).get("results", [])
                    current_page = data.get("articles", {}).get("page", 1)
                    total_pages = data.get("articles", {}).get("pages", 1)
                    all_articles.extend(articles)

                    print(f"✅ Page {current_page} of {total_pages} fetched. Total articles: {len(all_articles)}")

                    if current_page >= 1:
                        print("🎉 All pages have been fetched.")
                        break

                    params["articlesPage"] += 1
                    time.sleep(2)  # رعایت ریت‌لیمیت
                else:
                    print(f"❌ Error fetching page {params['articlesPage']}: HTTP status code {response.status_code}")
                    break

            except Exception as e:
                print(f"⚠️ Unexpected error on page {params['articlesPage']}: {e}")
                break

        with transaction.atomic():
            kb_objects = []

            for article in all_articles:
                uri = article.get('uri')
                title = article.get('title')
                body = article.get('body')
                url = article.get('url')

                # جلوگیری از ذخیره مقالات تکراری بر اساس uri
                if KnowledgeBase.objects.filter(uri=uri).exists():
                    continue

                # اطلاعات منبع
                source_data = article.get('source', {})
                source_title = source_data.get('title')
                source_uri = source_data.get('uri')
                source_data_type = source_data.get('dataType')

                # گرفتن یا ساخت منبع
                source_obj, _ = Source.objects.get_or_create(
                    title=source_title,
                    defaults={
                        "category": "real",
                        "source_uri": source_uri,
                        "source_data_type": source_data_type
                    }
                )

                # تبدیل تاریخ انتشار
                date_time_pub = None
                try:
                    date_time_pub = datetime.fromisoformat(article.get("dateTimePub").replace("Z", "+00:00"))
                except:
                    pass

                kb = KnowledgeBase(
                    title=title,
                    body=body,
                    url=url,
                    source=source_obj,
                    category="real",
                    image=article.get("image"),
                    uri=uri,
                    lang=article.get("lang"),
                    is_duplicate=article.get("isDuplicate"),
                    data_type=article.get("dataType"),
                    sim=article.get("sim"),
                    sentiment=article.get("sentiment"),
                    wgt=article.get("wgt"),
                    relevance=article.get("relevance"),
                    authors=article.get("authors"),
                    date_time_pub=date_time_pub
                )

                kb_objects.append(kb)

            # ذخیره همه اخبار به صورت bulk
            # saved_kbs = KnowledgeBase.objects.bulk_create(kb_objects)

            # we dont need it with autoLabeling
            # assign_default_labels_to_kbs(saved_kbs)

            # create_kb_process_status(saved_kbs)
            
        # print("count of records insert", len(saved_kbs))
        # celery BEAT and celery worker should execute concurrent
        return Response(
            {
                "data": all_articles,
                # "message": f"{len(saved_kbs)} news add to DB."
            },
            status=status.HTTP_200_OK
            )
    
        
       
class UpdateUnprocessedKBView(APIView):
    def post(self, request):

        # input_date = date(2025, 7, 6)
        # kbs = KnowledgeBase.objects.filter(
        #     Q(created_at__date=input_date) & Q(processed=False)
        # ) [:30]

        now = datetime.now()
        yesterday = now - timedelta(days=1)
        kbs = KnowledgeBase.objects.filter(Q(created_at__gte=yesterday) & Q(processed=False))[:2]
        print("sssssssssssss", len(kbs))

        count = 0
        # for kb in kbs:
            # send_kb_to_ai.delay(kb_id=kb.id)


            # print("aaaaaaaaaaaaaaaaaaaa", kb.id, kb.category)

            # headers = {
            # 'sahaa-ai-api': 'WGhgR5dOAEc34MI0Zpi5C2Y3LyjwT9Ex',
            # 'Content-Type': 'application/json',
            # }

            # payload = {
            #     'category': kb.category,
            #     'id': str(kb.id),
            #     'body': kb.body,
            # }
            # response = requests.post('http://89.42.199.251:5682/text/kb/add_news', params=payload, headers=headers, timeout=(3, 60))
            # print("2222222222222222222222", response.status_code)

            # if response.status_code == 200:
            #     kb.processed = True
            #     kb.save()
            #     print("✅ API sent for kb.id =", kb.id)
            # # return
            # count += 1
            # continue

        return Response(
            {"message": f"{count} task(s) sent to Celery."},
            status=status.HTTP_200_OK
    )
    
    

class ImportTestNewsContentExcelView(APIView):
    parser_classes = [MultiPartParser]

    def post(self, request):
        file_obj = request.FILES.get('file')
        if not file_obj:
            return Response({"error": "Excel file is required."}, status=400)

        wb = openpyxl.load_workbook(filename=BytesIO(file_obj.read()))
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        headers = rows[0]
        data_rows = rows[1:]

        title_idx = headers.index("title")
        body_idx = headers.index("body")
        social_idx = headers.index("social_media")
        date_time_pub_idx = headers.index("date_time_pub")

        kb_objects = []
        with transaction.atomic():
            for row in data_rows:
                batch_id = uuid4()
                title = row[title_idx]
                body = row[body_idx]
                social_title = row[social_idx]
                date_time_pub = row[date_time_pub_idx]


                if not (title and body and social_title):
                    continue

                # حذف دابل کوتیشن‌های اول و آخر فقط اگه هست
                if title and title.startswith('"') and title.endswith('"'):
                    title = title[1:-1]

                if body and body.startswith('"') and body.endswith('"'):
                    body = body[1:-1]

                existing_sm = SocialMedia.objects.filter(title=social_title).first()
                if existing_sm:
                    source_obj = existing_sm
                else:
                    source_obj = SocialMedia.objects.create(title=social_title)

                kb = KnowledgeBase(
                            title=title,
                            body=body,
                            social_media=source_obj,
                            import_batch_id=batch_id,
                            date_time_pub=date_time_pub,
                            percentages=None
                        )
                
                kb_objects.append(kb)

        # ذخیره همه اخبار به صورت bulk
        saved_kbs = KnowledgeBase.objects.bulk_create(kb_objects)
        print("✅count of records insert", len(saved_kbs))

        return Response({
            "message": "File imported successfully. Data is being processed asynchronously.",
            },
            status=status.HTTP_202_ACCEPTED
            )
    
class CheckNewsContentView(APIView):
    def get(self, request):
        # batch_id = request.query_params.get('batch_id')
        # if not batch_id:
        #     return Response({"error": "batch_id is required"}, status=400)
        # social_medias = SocialMedia.objects.filter(knowledgebase__import_batch_id=batch_id).distinct()
        paginator = CustomPagination()
        social_medias = SocialMedia.objects.all().distinct()


        result = []
        for sm in social_medias:
            counts = KnowledgeBaseLabelUser.objects.filter(
                knowledge_base__social_media=sm, 
                knowledge_base__import_batch_id__isnull=False,
            ) \
            .values('label__name') \
            .annotate(count=Count('id'))

            total = sum(item['count'] for item in counts)
            if total == 0:
                continue

            # محاسبه آمار مربوط به همین شبکه اجتماعی
            total_count = KnowledgeBase.objects.filter(
                social_media=sm,
                import_batch_id__isnull=False
            ).count()

            processed_count = KnowledgeBase.objects.filter(
                social_media=sm,
                import_batch_id__isnull=False,
                processed=True
            ).count()

            remaining_count = total_count - processed_count
            
            serializer = SocialMediaProcessByAiSerializer(sm,context={'request': self.request})
            
            result.append({
                "data": serializer.data,
                "percentages": {
                    item['label__name']: round((item['count'] / total) * 100, 2)
                    for item in counts
                },
                "stats": {
                    "total_count": total_count,
                    "processed_count": processed_count,
                    "remaining_count": remaining_count
                }
            })

       
        paginated = paginator.paginate_queryset(result, request)
        # return Response({
        #         "count": paginator.page.paginator.count,
        #         "next": paginator.get_next_link(),
        #         "previous": paginator.get_previous_link(),
        #         "results": paginated
        #     }, status=status.HTTP_200_OK)
        return Response({
            "result": result,
        }, status=status.HTTP_200_OK)


class ImportNewsContentDetailView(APIView):
    def get(self, request):
        # batch_id = request.query_params.get('batch_id')
        # if not batch_id:
            # return Response({"error": "batch_id is required"}, status=400)
        paginator = CustomPagination()
        
        social_media_id = request.query_params.get('social_media_id')
        if not social_media_id:
            return Response({"error": "social_media_id is required"}, status=400)

        kb_items = KnowledgeBase.objects.filter(
            import_batch_id__isnull=False,
            social_media_id=social_media_id,
        ).select_related('social_media')

        data = []
        for kb in kb_items:
            # kb_label_user = KnowledgeBaseLabelUser.objects.filter(
            #     knowledge_base=kb
            # ).select_related('label').first()
            
            # label = kb_label_user.label.name if kb_label_user else None

            serializer = KnowledgeBaseProcessByAiSerializer(kb ,context={'request': self.request})
            # data.append({
            #     "title": kb.title,
            #     "body": kb.body,
            #     "Knowledge_base_id": kb.id,
            #     "social_media_id": kb.social_media.id,
            #     "social_media": kb.social_media.title,
            #     "label_id": kb_label_user.label.id,
            #     "label": label,
            #     "created_at": kb.created_at.strftime('%Y-%m-%d %H:%M:%S'),
            #     "updated_at": kb.updated_at.strftime('%Y-%m-%d %H:%M:%S'),
            # })

            data.append({
                "knowledge_base_item": serializer.data,
            })


        paginated = paginator.paginate_queryset(data, request)
        # return Response({
        #         "count": paginator.page.paginator.count,
        #         "next": paginator.get_next_link(),
        #         "previous": paginator.get_previous_link(),
        #         "results": paginated
        #     }, status=status.HTTP_200_OK)
        return Response(data, status=status.HTTP_200_OK)
    
    
class DownloadKnowledgeBaseSourceType(APIView):
    # permission_classes = [IsAuthenticated]

    # def get(self, request):
    #     # ساخت فایل CSV با encoding مناسب برای فارسی
    #     temp = tempfile.NamedTemporaryFile(delete=False, suffix='.csv', mode='w', newline='', encoding='utf-8-sig')
    #     writer = csv.writer(temp)

    #     # نوشتن عنوان ستون‌ها
    #     writer.writerow([
    #         'id', 'title', 'body', 'url', 'date_time_pub',
    #         'source_id', 'source__title', 'uri', 'is_news',
    #         'sim', 'sentiment', 'wgt', 'relevance', 'authors'
    #     ])

    #     # کوئری دیتابیس
    #     # queryset = KnowledgeBase.objects.select_related('source').filter(source__isnull=False)[:1000]
    #     queryset = KnowledgeBase.objects.select_related('source').filter(source__isnull=False, is_news=True)


    #     def clean_text(text):
    #         return text.replace('\n', ' ').replace('\r', ' ').replace('\t', ' ') if text else ''
    #     for kb in queryset:
    #         writer.writerow([
    #             str(kb.id),
    #             clean_text(str(kb.title)),
    #             clean_text(str(kb.body)),
    #             str(kb.url),
    #             jdatetime.datetime.fromgregorian(datetime=kb.date_time_pub).strftime('%Y-%m-%d') if kb.date_time_pub else '',
    #             str(kb.source.id) if kb.source else '',
    #             clean_text(str(kb.source.title)) if kb.source else '',
    #             str(kb.uri),
    #             str(kb.is_news),
    #             str(kb.sim),
    #             str(kb.sentiment),
    #             str(kb.wgt),
    #             str(kb.relevance),
    #             clean_text(str(kb.authors))
    #         ])

    #     temp.close()

    #     # باز کردن فایل و ارسال آن
    #     f = open(temp.name, 'rb')
    #     response = FileResponse(f, content_type='text/csv')
    #     response['Content-Disposition'] = 'attachment; filename="knowledgebase.csv"'

    #     return response
    
    # this function get excel files from data in chunk of 1000 record in fromat of ZIP 
    def safe_excel_value(self, value):
        if value is None:
            return ''
        if isinstance(value, list):
            return ", ".join(str(v) for v in value)
        return str(value)
    
    def get(self, request):
        # تعداد رکورد در هر فایل اکسل
        chunk_size = 1000

        # all_data = list(
        #     KnowledgeBase.objects.select_related('source')
        #     .filter(source__isnull=False, is_news=True)
        #     .values(
        #         'id', 'title', 'body', 'url', 'date_time_pub',
        #         'source_id','source__title','source__source_uri' ,'uri', 'is_news',
        #         'sim', 'sentiment', 'wgt', 'relevance', 'authors'
        #     )
        # )

        all_data = list(
            KnowledgeBase.objects.select_related('source')
            .filter(source__isnull=False)
            .values(
                'id', 'title', 'body', 'url', 'date_time_pub',
                'source_id','source__title','source__source_uri' ,'uri', 'is_news',
                'sim', 'sentiment', 'wgt', 'relevance', 'authors'
            )
        )
        
        headers = [
            'id', 'title', 'body', 'url', 'date_time_pub',
            'source_id','source_title','source_uri' ,'uri', 'is_news',
            'sim', 'sentiment', 'wgt', 'relevance', 'authors'
        ]

        # ایجاد فایل zip در حافظه
        zip_buffer = BytesIO()
        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:

            # تقسیم داده‌ها به بخش‌های chunk_size
            for i in range(0, len(all_data), chunk_size):
                chunk = all_data[i:i + chunk_size]
                
                # ساخت فایل اکسل در حافظه برای هر بخش
                excel_buffer = BytesIO()
                workbook = xlsxwriter.Workbook(excel_buffer)
                worksheet = workbook.add_worksheet('KnowledgeBase')

                # فرمت‌ها
                header_format = workbook.add_format({
                    'bold': True,
                    'bg_color': '#D9E1F2',
                    'border': 1,
                    'align': 'center',
                    'valign': 'vcenter'
                })
                cell_format = workbook.add_format({
                    'align': 'center',
                    'valign': 'vcenter',
                    'text_wrap': True
                })

                # تنظیم عرض ستون‌ها
                widths = [15, 15, 50, 25, 15, 15, 15, 15, 15, 15, 15, 15, 15, 15, 25]
                for col, width in enumerate(widths):
                    worksheet.set_column(col, col, width)

                # نوشتن هدرها
                for col, header in enumerate(headers):
                    worksheet.write(0, col, header, header_format)

                # نوشتن داده‌ها
                for row, item in enumerate(chunk, start=1):
                    row_data = [
                        self.safe_excel_value(item['id']),
                        self.safe_excel_value(item['title']),
                        self.safe_excel_value((item['body'] or '')[:1000]),
                        self.safe_excel_value(item['url']),
                        jdatetime.datetime.fromgregorian(datetime=item['date_time_pub']).strftime('%Y-%m-%d') if item['date_time_pub'] else '',
                        self.safe_excel_value(item['source_id']),
                        self.safe_excel_value(item['source__title']),
                        self.safe_excel_value(item['source__source_uri']),
                        self.safe_excel_value(item['uri']),
                        self.safe_excel_value(item['is_news']),
                        self.safe_excel_value(item['sim']),
                        self.safe_excel_value(item['sentiment']),
                        self.safe_excel_value(item['wgt']),
                        self.safe_excel_value(item['relevance']),
                        self.safe_excel_value(item['authors']),
                    ]
                    for col, value in enumerate(row_data):
                        worksheet.write(row, col, value, cell_format)

                # فریز کردن ردیف اول و افزودن autofilter
                worksheet.freeze_panes(1, 0)
                worksheet.autofilter(0, 0, len(chunk), len(headers) - 1)

                workbook.close()
                excel_buffer.seek(0)

                # اضافه کردن فایل اکسل به فایل زیپ
                filename = f"KnowledgeBase_part_{i//chunk_size + 1}.xlsx"
                zip_file.writestr(filename, excel_buffer.read())

        zip_buffer.seek(0)

        response = HttpResponse(zip_buffer, content_type='application/zip')
        response['Content-Disposition'] = 'attachment; filename=KnowledgeBase_export.zip'
        return response

from django.core.files.storage import FileSystemStorage
class ImportLabelsViaExcel(APIView):
    def post(self, request):
        excel_file = request.FILES.get("file")
        if not excel_file:
            return Response({"error": "هیچ فایلی ارسال نشده"}, status=400)

        try:
                        # ذخیره موقت فایل
            # fs = FileSystemStorage()
            # filename = fs.save(excel_file.name, excel_file)
            # file_path = fs.path(filename)

            # باز کردن اکسل
            wb = openpyxl.load_workbook(filename=BytesIO(excel_file.read()), data_only=True)

            # انتخاب sheet خاص
            if "FinalDataset" not in wb.sheetnames:
                return Response({"error": "no sheet named FinalDataset not found"}, status=400)

            sheet = wb["FinalDataset"]

            # گرفتن header ها
            headers = [cell.value for cell in sheet[1]]

            kb_idx = headers.index("ID") if "ID" in headers else None
            label_idx = headers.index("Label") if "Label" in headers else None

            if kb_idx is None or label_idx is None:
                return Response({"error": "فایل اکسل باید شامل ستون‌های 'ID' و 'Label' باشد"}, status=400)

            # کاربر سیستمی
            system_user, _  = User.objects.get_or_create(
            email="system_user@yourapp.com",
            defaults={"username": "system_user", "name": "System User", "password": "system_user"}
            )
            # مپ لیبل‌ها
            label_map = {
                "حقیقت": 1,
                "فریب دهی (اطلاعات نادرست عمدی / گمراه‌کننده)": 2,
            }

            imported_count = 0
            skipped_count = 0

            for row in sheet.iter_rows(min_row=2, values_only=True):
                kb_id = row[kb_idx]
                label_name = row[label_idx]

                if not kb_id or not label_name:
                    skipped_count += 1
                    print("ssssssssssss", kb_id, label_name)
                    continue
                label_id = label_map.get(label_name)
                if not label_id:
                    print("cccccccccccc", label_name)
                    skipped_count += 1
                    continue
               
                KnowledgeBaseLabelUser.objects.get_or_create(
                    knowledge_base_id=kb_id,
                    label_id=label_id,
                    user=system_user
                )
                imported_count += 1

            return Response({
                "message": "",
                "imported": imported_count,
                "skipped": skipped_count
            })

        except Exception as e:
            return Response({"error": str(e)}, status=500)
